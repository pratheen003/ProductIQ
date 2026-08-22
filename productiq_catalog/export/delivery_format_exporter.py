"""
ProductIQ Catalog Delivery Format Exporter
===========================================
Generates exact-header delivery format exports (.xlsx and .csv) matching the
exact 252-column header sequence of Unihack__Expected_Output_-_Delivery_Format.csv.
Unpopulated columns remain genuinely blank to strictly uphold the no-fabrication rule.
"""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from productiq_catalog.extraction.input_loader import InputDatasetLoader
from productiq_catalog.enrichment.catalog_enricher import CatalogPipeline
from productiq_catalog.ground_truth.ingest import GroundTruthStore
from productiq_catalog.schema.models import CatalogProduct


class DeliveryFormatExporter:
    """
    Exports enriched catalog data into the exact 252-column delivery format schema.
    """

    def __init__(
        self,
        ground_truth_csv_path: Optional[Path | str] = None,
        output_dir: Optional[Path | str] = None,
    ):
        self.root = Path(__file__).resolve().parent.parent.parent
        self.gt_csv_path = (
            Path(ground_truth_csv_path)
            if ground_truth_csv_path
            else self.root / "data" / "catalog" / "ground_truth" / "Unihack__Expected_Output_-_Delivery_Format.csv"
        )
        self.output_dir = (
            Path(output_dir)
            if output_dir
            else self.root / "data" / "catalog" / "processed"
        )
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self._headers: List[str] = self._extract_canonical_headers()

    def _extract_canonical_headers(self) -> List[str]:
        """
        Extract the exact header row from the ground truth reference CSV.
        """
        if not self.gt_csv_path.exists():
            raise FileNotFoundError(f"Ground truth reference file not found at {self.gt_csv_path}")

        with open(self.gt_csv_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            headers = next(reader)

        return [h.strip() for h in headers]

    @property
    def headers(self) -> List[str]:
        return list(self._headers)

    @property
    def header_count(self) -> int:
        return len(self._headers)

    def map_product_to_row(self, product: CatalogProduct) -> Dict[str, Any]:
        """
        Map an enriched CatalogProduct to a dictionary keyed by the 252 canonical headers.
        Unpopulated fields remain genuinely empty (empty string).
        """
        row_dict: Dict[str, Any] = {h: "" for h in self._headers}

        # 1. Raw Input Columns
        if "Mfg_Part_Num" in row_dict:
            row_dict["Mfg_Part_Num"] = product.mfg_part_num or ""
        if "Part_Desc" in row_dict:
            row_dict["Part_Desc"] = product.part_desc or ""
        if "Part_Manuf" in row_dict:
            row_dict["Part_Manuf"] = product.raw_input.part_manuf or ""
        if "E1_Brand" in row_dict:
            row_dict["E1_Brand"] = product.raw_input.e1_brand or ""
        if "Unilog_Brand" in row_dict:
            row_dict["Unilog_Brand"] = product.raw_input.unilog_brand or ""
        if "DIB_Brand" in row_dict:
            row_dict["DIB_Brand"] = product.raw_input.dib_brand or ""

        # 2. Canonical Identity Fields
        if "MANUFACTURER_NAME" in row_dict:
            row_dict["MANUFACTURER_NAME"] = product.manufacturer_name.value or ""
        if "BRAND_NAME" in row_dict:
            row_dict["BRAND_NAME"] = product.brand_name.value or ""
        if "TRADE_NAME" in row_dict:
            row_dict["TRADE_NAME"] = product.trade_name.value or ""
        if "MANUFACTURER_PART_NUMBER" in row_dict:
            row_dict["MANUFACTURER_PART_NUMBER"] = product.manufacturer_part_number.value or product.mfg_part_num or ""
        if "Product Name" in row_dict:
            row_dict["Product Name"] = product.product_name.value or ""
        if "Classpath" in row_dict:
            row_dict["Classpath"] = product.classpath.value or ""
        if "SERIES" in row_dict:
            row_dict["SERIES"] = product.series.value or ""
        if "SHORT_DESCRIPTION" in row_dict:
            row_dict["SHORT_DESCRIPTION"] = product.short_desc.value or ""
        if "LONG_DESCRIPTION" in row_dict:
            row_dict["LONG_DESCRIPTION"] = product.long_desc.value or ""

        # 3. Attribute Triples (ATTRIBUTE_LABEL n, ATTRIBUTE_VALUE n, ATTRIBUTE_UOM n)
        for idx, attr in enumerate(product.attributes, start=1):
            if idx > 50:
                break
            # Ground truth format has a space e.g. "ATTRIBUTE_LABEL 1"
            label_col = f"ATTRIBUTE_LABEL {idx}"
            val_col = f"ATTRIBUTE_VALUE {idx}"
            uom_col = f"ATTRIBUTE_UOM {idx}"

            if label_col in row_dict:
                row_dict[label_col] = attr.label or ""
            elif f"ATTRIBUTE_LABEL_{idx}" in row_dict:
                row_dict[f"ATTRIBUTE_LABEL_{idx}"] = attr.label or ""

            if val_col in row_dict:
                row_dict[val_col] = str(attr.value) if attr.value is not None else ""
            elif f"ATTRIBUTE_VALUE_{idx}" in row_dict:
                row_dict[f"ATTRIBUTE_VALUE_{idx}"] = str(attr.value) if attr.value is not None else ""

            if uom_col in row_dict:
                row_dict[uom_col] = attr.uom or ""
            elif f"ATTRIBUTE_UOM_{idx}" in row_dict:
                row_dict[f"ATTRIBUTE_UOM_{idx}"] = attr.uom or ""

        return row_dict

    def export_all(
        self,
        pipeline: Optional[CatalogPipeline] = None,
        input_loader: Optional[InputDatasetLoader] = None,
    ) -> Dict[str, Any]:
        """
        Enrich all 1,000 input rows and export to both .xlsx and .csv files.
        """
        loader = input_loader or InputDatasetLoader()
        pipe = pipeline or CatalogPipeline()

        rows = loader.get_all()
        total_rows = len(rows)

        mapped_rows: List[Dict[str, Any]] = []
        for r in rows:
            product = pipe.process_row(r)
            mapped_rows.append(self.map_product_to_row(product))

        # 1. Write CSV
        csv_path = self.output_dir / "productiq_delivery_output.csv"
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=self._headers)
            writer.writeheader()
            writer.writerows(mapped_rows)

        # 2. Write Excel (.xlsx)
        xlsx_path = self.output_dir / "productiq_delivery_output.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Unilog Delivery Output"

        # Style header row
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

        ws.append(self._headers)
        for col_idx in range(1, len(self._headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Append data rows
        for row_dict in mapped_rows:
            row_values = [row_dict[h] for h in self._headers]
            ws.append(row_values)

        # Freeze header row
        ws.freeze_panes = "A2"
        wb.save(xlsx_path)

        return {
            "total_rows": total_rows,
            "total_columns": len(self._headers),
            "csv_path": str(csv_path),
            "xlsx_path": str(xlsx_path),
            "csv_size_bytes": csv_path.stat().st_size,
            "xlsx_size_bytes": xlsx_path.stat().st_size,
        }
