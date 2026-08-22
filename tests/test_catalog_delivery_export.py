"""
ProductIQ Catalog Exact-Header Delivery Format Export Test Suite
=================================================================
Tests that exported XLSX and CSV match the exact 252 ground-truth headers,
populated columns carry genuine values, unpopulated columns remain blank,
and the API download endpoint functions properly.
"""
from __future__ import annotations

import csv
from pathlib import Path
import openpyxl
import pytest
from fastapi.testclient import TestClient

from productiq_catalog.export.delivery_format_exporter import DeliveryFormatExporter
from productiq.api.app import app

client = TestClient(app)


class TestDeliveryFormatExporter:
    """Test delivery format export mechanics and schema compliance."""

    def test_canonical_header_count_and_order(self):
        exporter = DeliveryFormatExporter()
        gt_csv = Path(__file__).resolve().parent.parent / "data" / "catalog" / "ground_truth" / "Unihack__Expected_Output_-_Delivery_Format.csv"
        
        with open(gt_csv, "r", encoding="utf-8", errors="replace") as f:
            gt_headers = [h.strip() for h in next(csv.reader(f))]

        assert len(gt_headers) == 252
        assert exporter.header_count == 252
        assert exporter.headers == gt_headers

    def test_export_all_generates_valid_files(self):
        exporter = DeliveryFormatExporter()
        result = exporter.export_all()

        assert result["total_rows"] == 1000
        assert result["total_columns"] == 252

        xlsx_path = Path(result["xlsx_path"])
        csv_path = Path(result["csv_path"])

        assert xlsx_path.exists()
        assert csv_path.exists()
        assert xlsx_path.stat().st_size > 100000
        assert csv_path.stat().st_size > 100000

    def test_exported_csv_header_and_row_fidelity(self):
        csv_path = Path(__file__).resolve().parent.parent / "data" / "catalog" / "processed" / "productiq_delivery_output.csv"
        assert csv_path.exists()

        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader)
            rows = list(reader)

        assert len(headers) == 252
        assert len(rows) == 1000

        # Spot check row 0 (DCB518ASTS06G)
        row_0 = dict(zip(headers, rows[0]))
        assert row_0["Mfg_Part_Num"] == "DCB518ASTS06G"
        assert row_0["Part_Desc"] == "DCB518ASTS06G Diablo 1/2\"x18\" - Sanding Belt 6pc"
        assert row_0["Product Name"] == "Sanding Belt"
        # Unpopulated columns must be empty strings
        assert row_0["MFR URL"] == ""
        assert row_0["Ref URL 1"] == ""
        assert row_0["Video Link"] == ""

    def test_exported_xlsx_structure_and_no_corruption(self):
        xlsx_path = Path(__file__).resolve().parent.parent / "data" / "catalog" / "processed" / "productiq_delivery_output.xlsx"
        assert xlsx_path.exists()

        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        assert ws.title == "Unilog Delivery Output"
        assert ws.max_row == 1001

        row_iter = ws.iter_rows(values_only=True)
        headers = list(next(row_iter))
        assert len(headers) == 252

        first_data_row = dict(zip(headers, next(row_iter)))
        assert first_data_row["Mfg_Part_Num"] == "DCB518ASTS06G"
        # Unpopulated cells must be None in openpyxl
        assert first_data_row["MFR URL"] is None or first_data_row["MFR URL"] == ""

        wb.close()


class TestDeliveryFormatAPIEndpoint:
    """Test API export/download route."""

    def test_export_xlsx_endpoint(self):
        resp = client.get("/api/catalog/export/delivery-format?format=xlsx")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert "attachment" in resp.headers["content-disposition"]
        assert "productiq_delivery_output.xlsx" in resp.headers["content-disposition"]
        assert len(resp.content) > 100000

    def test_export_csv_endpoint(self):
        resp = client.get("/api/catalog/export/delivery-format?format=csv")
        assert resp.status_code == 200
        assert "text/csv" in resp.headers["content-type"]
        assert "attachment" in resp.headers["content-disposition"]
        assert "productiq_delivery_output.csv" in resp.headers["content-disposition"]
        assert len(resp.content) > 100000
